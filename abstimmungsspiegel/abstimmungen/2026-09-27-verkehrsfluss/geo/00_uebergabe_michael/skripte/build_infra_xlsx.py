#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
R=json.load(open('infra_results.json'))
FACS=json.load(open('facilities.json'))
def canon(g): return (g or '').replace(' (SH)','').strip()
DS=Counter(f['cat'] for f in FACS)
ADM=defaultdict(Counter)
for f in FACS: ADM[canon(f['gem'])][f['cat']]+=1
FONT='Calibri'
hdr=PatternFill('solid',fgColor='2F3B52'); band=PatternFill('solid',fgColor='F2F3F5'); totf=PatternFill('solid',fgColor='E7EAF0')
thin=Side(style='thin',color='D9D9D9'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def km(m): return round(m/1000,2)
def pct(x,t): return round(100*x/t) if t else 0
TOT=sum(r['stats']['betroffen'] for r in R)
def ctot(k): return sum(r['stats'][k] for r in R)
CATS=[('schule','S'),('kindergarten','K'),('altersheim','H'),('sozial','So')]
wb=Workbook()

ws=wb.active; ws.title='Uebersicht'; ws.sheet_view.showGridLines=False
ws.merge_cells('A1:I1'); c=ws.cell(1,1,'Betroffene Kantonsstrassen und Nähe zu sensiblen Nutzungen (100 / 300 / 500 m), je Gemeinde')
c.font=Font(name=FONT,bold=True,size=13,color='2F3B52')
ws.merge_cells('A2:I2'); c=ws.cell(2,1,'Betroffen = verkehrsorientierte Kantonsstrassen innerorts. Kandidat kurzer Abschnitt = Strasse innerhalb 100 m (eng) bzw. 300 m einer Nutzung. Umkreis = Luftlinie. LV95.')
c.font=Font(name=FONT,size=9,italic=True,color='5A6472'); c.alignment=Alignment(wrap_text=True); ws.row_dimensions[2].height=28
H=['Gemeinde','betroffen km','≤100 m km','≤100 m %','≤300 m km','≤300 m %','≤500 m km','≤500 m %','Anlagen S/K/H/So']
h0=4
for j,x in enumerate(H,1):
    cc=ws.cell(h0,j,x); cc.font=Font(name=FONT,bold=True,size=9,color='FFFFFF'); cc.fill=hdr; cc.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cc.border=bd
ws.row_dimensions[h0].height=40
rr=h0+1
for i,r in enumerate(R):
    s=r['stats']; a=ADM[canon(r['gemeinde'])]
    vals=[r['gemeinde'].replace(' (SH)',''),km(s['betroffen']),km(s['alle_100']),pct(s['alle_100'],s['betroffen']),
          km(s['alle_300']),pct(s['alle_300'],s['betroffen']),km(s['alle_500']),pct(s['alle_500'],s['betroffen']),
          f"{a['schule']}/{a['kindergarten']}/{a['altersheim']}/{a['sozial']}"]
    for j,v in enumerate(vals,1):
        cc=ws.cell(rr,j,v); cc.font=Font(name=FONT,size=9); cc.border=bd; cc.alignment=Alignment(horizontal='left' if j==1 else 'center')
        if j in (2,3,5,7): cc.number_format='0.00'
        if j in (4,6,8): cc.number_format='0'
    if i%2==1:
        for j in range(1,10): ws.cell(rr,j).fill=band
    rr+=1
ST={c:sum(ADM[canon(r['gemeinde'])][c] for r in R) for c,_ in CATS}
tot=['Kanton total',km(TOT),km(ctot('alle_100')),pct(ctot('alle_100'),TOT),km(ctot('alle_300')),pct(ctot('alle_300'),TOT),
     km(ctot('alle_500')),pct(ctot('alle_500'),TOT),f"{ST['schule']}/{ST['kindergarten']}/{ST['altersheim']}/{ST['sozial']}"]
for j,v in enumerate(tot,1):
    cc=ws.cell(rr,j,v); cc.font=Font(name=FONT,bold=True,size=9); cc.fill=totf; cc.border=bd; cc.alignment=Alignment(horizontal='left' if j==1 else 'center')
    if j in (2,3,5,7): cc.number_format='0.00'
    if j in (4,6,8): cc.number_format='0'
for j,w in zip(range(1,10),[20,13,12,10,12,10,12,10,17]): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes='A5'
ws.cell(rr+2,1,f'Datensatz kantonsweit: {len(FACS)} Standorte ({DS["schule"]} Schulen, {DS["kindergarten"]} Kindergärten, {DS["altersheim"]} Alters-/Pflegeheime, {DS["sozial"]} weitere Sozialeinrichtungen).').font=Font(name=FONT,size=8,italic=True,color='5A6472')

gs=wb.create_sheet('Nach_Gruppe'); gs.sheet_view.showGridLines=False
gs.merge_cells('A1:M1'); c=gs.cell(1,1,'Betroffene Strasse innerhalb 100 / 300 / 500 m je Gruppe (km), je Gemeinde')
c.font=Font(name=FONT,bold=True,size=12,color='2F3B52')
H2=['Gemeinde']
for _,ab in CATS:
    for r_ in (100,300,500): H2.append(f'{ab} ≤{r_}m')
for j,x in enumerate(H2,1):
    cc=gs.cell(3,j,x); cc.font=Font(name=FONT,bold=True,size=8,color='FFFFFF'); cc.fill=hdr; cc.alignment=Alignment(horizontal='center',wrap_text=True); cc.border=bd
gs.row_dimensions[3].height=26
r2=4
for r in R:
    s=r['stats']; gs.cell(r2,1,r['gemeinde'].replace(' (SH)','')).font=Font(name=FONT,size=9); gs.cell(r2,1).border=bd
    j=2
    for cat,_ in CATS:
        for rad in (100,300,500):
            cc=gs.cell(r2,j,km(s[f'{cat}_{rad}'])); cc.font=Font(name=FONT,size=9); cc.border=bd; cc.number_format='0.00'; cc.alignment=Alignment(horizontal='center'); j+=1
    r2+=1
gs.column_dimensions['A'].width=20
for j in range(2,14): gs.column_dimensions[get_column_letter(j)].width=8
gs.freeze_panes='B4'

es=wb.create_sheet('Kandidaten_Beispiele'); es.sheet_view.showGridLines=False
es.merge_cells('A1:E1'); c=es.cell(1,1,'Kandidaten kurzer Abschnitte je Gemeinde: betroffene Strassen innerhalb 100 m (eng) und 300 m einer Nutzung, benannt')
c.font=Font(name=FONT,bold=True,size=12,color='2F3B52')
for j,x in enumerate(['Gemeinde','Umkreis','Strasse (Name aus Netzbelastung)','Länge (m)','Anteil an Kandidaten (Umkreis, Gemeinde)'],1):
    cc=es.cell(3,j,x); cc.font=Font(name=FONT,bold=True,size=9,color='FFFFFF'); cc.fill=hdr; cc.border=bd; cc.alignment=Alignment(horizontal='center',wrap_text=True)
es.row_dimensions[3].height=30
r3=4
for r in R:
    for rad,names,length in [(100,r['cand100_names'],r['cand100_len']),(300,r['cand300_names'],r['cand300_len'])]:
        if not names:
            es.cell(r3,1,r['gemeinde'].replace(' (SH)','')).font=Font(name=FONT,size=9); es.cell(r3,1).border=bd
            es.cell(r3,2,f'≤{rad} m').font=Font(name=FONT,size=9); es.cell(r3,2).border=bd
            cc=es.cell(r3,3,'— keine —'); cc.font=Font(name=FONT,size=9,italic=True,color='888888'); cc.border=bd
            es.cell(r3,4,0).border=bd; es.cell(r3,5,'').border=bd; r3+=1; continue
        for k,(nm,l) in enumerate(names):
            nm2='unbenannt' if nm in ('Namenlos','','unbenannt') else nm
            es.cell(r3,1,r['gemeinde'].replace(' (SH)','') if k==0 else '').font=Font(name=FONT,size=9,bold=(k==0)); es.cell(r3,1).border=bd
            es.cell(r3,2,f'≤{rad} m' if k==0 else '').font=Font(name=FONT,size=9); es.cell(r3,2).border=bd
            es.cell(r3,3,nm2).font=Font(name=FONT,size=9); es.cell(r3,3).border=bd
            cc=es.cell(r3,4,round(l)); cc.font=Font(name=FONT,size=9); cc.border=bd; cc.number_format='0'; cc.alignment=Alignment(horizontal='center')
            cc=es.cell(r3,5,round(100*l/length)/100 if length else 0); cc.number_format='0%'; cc.font=Font(name=FONT,size=9); cc.border=bd; cc.alignment=Alignment(horizontal='center')
            r3+=1
for j,w in zip(range(1,6),[18,10,40,12,22]): es.column_dimensions[get_column_letter(j)].width=w
es.freeze_panes='A4'
wb.save('Auswertung_Infrastruktur_kurze_Abschnitte_SH.xlsx')
print('xlsx written')
