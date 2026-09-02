#!/usr/bin/env python3
"""
Kantonsrat Schaffhausen - Abstimmungs-Scraper
==============================================
Findet automatisch alle Sitzungen auf der Übersichtsseite, lädt die
"Abstimmungsergebnisse"-Excel-Dateien herunter, parst sie und schreibt
ein konsolidiertes all_sessions.json für das Dashboard.

Ausführen:  python3 scraper.py
Abhängigkeiten:  openpyxl, requests
    pip install openpyxl requests

Kein Browser nötig: die Sitzungsübersicht auf sh.ch lädt ihre Kacheln über
den öffentlichen Listen-Endpoint /CMS/lists/list nach. Die Kachelansicht der
Webseite zeigt pro Jahr nur 12 Einträge, der Endpoint liefert dagegen alle.
Genau daran lag es, dass früher die ersten Sitzungen eines Jahres fehlten.
"""

import re, json, sys, time, html, hashlib, threading
import concurrent.futures as cf
from pathlib import Path
import requests
import openpyxl

# --- Konfiguration ---
UEBERSICHT_URL = ("https://sh.ch/CMS/Webseite/Kanton-Schaffhausen/Beh-rde/"
                  "Parlament/Der-Kantonsrat/Sitzungen-des-Kantonsrats---2274767-DE.html")
BASE = "https://sh.ch"
ROOT = Path(__file__).resolve().parent.parent    # Projektwurzel (scripts/ liegt darunter)
DATA = ROOT / "data"
OUT_DIR = DATA / "raw"           # heruntergeladene xlsx
OUT_DIR.mkdir(parents=True, exist_ok=True)
PROT_DIR = DATA / "protokolle"   # heruntergeladene Wortprotokoll-PDFs
PROT_DIR.mkdir(parents=True, exist_ok=True)
JSON_OUT = DATA / "all_sessions.json"            # konsolidierte Daten fürs Dashboard
INDEX_JSON = DATA / "sh_index.json"              # Cache: cid -> Titel, Datum, Dateien
WORKERS = 8                                      # parallele Abfragen gegen sh.ch

# Frühestes Jahr, das berücksichtigt wird. Elektronische (namentliche)
# Abstimmungen gibt es im Kantonsrat Schaffhausen seit 2018; ältere Sitzungen
# haben keine Abstimmungs-Excel und werden ohnehin übersprungen.
AB_JAHR = 2018

# Amtsdauern des Kantonsrats: vier Jahre, Beginn am 1. Januar nach den
# Gesamterneuerungswahlen im September davor. Nummer -> (von, bis) als
# (Jahr, Monat, Tag). Neue Legislaturen hier ergänzen.
LEGISLATUREN = {
    1: ((2017, 1, 1), (2020, 12, 31)),
    2: ((2021, 1, 1), (2024, 12, 31)),
    3: ((2025, 1, 1), (2028, 12, 31)),
}

# Listen-Endpoint der Sitzungsübersicht (liefert die Kacheln als HTML-Fragment).
# Die Parameter stammen aus dem AJAX-Aufruf der Webseite; rows/start blättern.
LIST_URL = "https://sh.ch/CMS/lists/list"
CONTENT_URL = "https://sh.ch/CMS/content"      # JSON je Inhaltsobjekt
LIST_PARAMS = {
    "sort": "sortable_datetime desc",
    "filter_customposttypeid_int": "403",          # Inhaltstyp "Event"
    "filter_published_string": "published",
    "filter_approvedpaths_string": "*/1752/8540/1753/1765/1755/1763/1805/40608/*",
    "filter_text": "()",
    "filter_language_string": "DE",
    "status": "",
    "kioskid": "be0e2565-f587-3475-8be2-1ef3b1e18133",
    "mode": "list",
    "append": "true",
    "slider": "false",
    "language": "DE",
    "domainpath": "/1752/8540/1753/",
    "kiosktype": "kioskWidget",
}
LIST_ROWS = 100          # der Server deckelt bei 100 Treffern pro Aufruf
# Achtung: sh.ch liefert je nach User-Agent eine abgespeckte Seite ohne die
# eingebetteten Dateiangaben. Mit browserüblichen Kopfzeilen kommt die volle
# Seite. Ohne das fehlen auf vielen Detailseiten sämtliche Downloads.
UA = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "de-CH,de;q=0.9",
}

# Dateien, die auf jeder Seite im Seitengerüst stecken und keine Sitzungsdateien sind
SEITEN_DATEIEN = ("favicon.ico", "apple-touch-icon.png", "vorlage_shch_icon.jpg",
                  "sh_ch_logo_farbe.svg")

# Abstimmungsdateien erkennen. sh.ch hat die Benennung über die Jahre dreimal
# gewechselt, darum mehrere Muster:
#   Abst_2018-07.xlsx                          (2018)
#   20190617-Vormittag-Kantonsrat.xlsx         (2018 bis 2022)
#   20230925-Vormittag.xlsx                    (2022 bis 2023)
#   20260824 Abstimmungsergebnisse Vormittag.xlsx  (seit Ende 2023)
XLSX_MUSTER = (
    re.compile(r"abstimmun[gs]*e?rgebnis", re.I),   # inkl. Tippfehler "Abstimmungergebnisse"
    re.compile(r"^\s*20\d{6}[-_\s]"),
    re.compile(r"^\s*abst[_\s-]?20\d{2}", re.I),
)
# Dateien, die trotz passendem Muster keine Abstimmungsergebnisse sind
XLSX_AUSSCHLUSS = ("traktand", "beschluss", "krb ", "protokoll", "einladung", "rangliste")
# Wortprotokoll erkennen: der Link-Text enthält "protokoll" oder folgt dem Muster
# "<Nr>. Sitzung vom <Datum>.pdf" (so heissen die Protokolle aktuell auf sh.ch).
# Ausgeschlossen werden Beschlüsse (KRB), Abstimmungsergebnisse und Traktanden.
PROT_EXCLUDE = ("abstimmungsergebnis", "krb", "traktand", "beschluss")


# ---------------------------------------------------------------------------
# SCHRITT 1: Sitzungen + Datei-Links von der Übersichtsseite einsammeln
# ---------------------------------------------------------------------------
DETAIL_URL_TMPL = ("https://sh.ch/CMS/Webseite/Kanton-Schaffhausen/Beh-rde/"
                   "Parlament/Der-Kantonsrat-{cid}-DE.html")


def _session_ids():
    """Alle Sitzungs-Kacheln (contentid) über den Listen-Endpoint einsammeln.

    Der Endpoint liefert maximal 100 Treffer pro Aufruf, darum wird mit
    start=0,100,... geblättert, bis nichts Neues mehr kommt."""
    s = requests.Session(); s.headers.update(UA)
    s.headers["X-Requested-With"] = "XMLHttpRequest"
    ids, start = [], 0
    while True:
        p = dict(LIST_PARAMS, rows=str(LIST_ROWS), start=str(start))
        r = s.get(LIST_URL, params=p, timeout=60); r.raise_for_status()
        gefunden = re.findall(r'contentid="(\d+)"', r.text)
        neu = [c for c in dict.fromkeys(gefunden) if c not in ids]
        if not neu:
            break
        ids += neu
        start += LIST_ROWS
        time.sleep(0.3)
    return ids, s


def _content(sess, cid):
    """Ein Inhaltsobjekt über die JSON-Schnittstelle des CMS holen."""
    r = sess.get(CONTENT_URL, params={"contentid": str(cid), "language": "DE"},
                 timeout=60, headers={"Accept": "application/json, text/javascript, */*; q=0.01",
                                      "X-Requested-With": "XMLHttpRequest"})
    r.raise_for_status()
    return r.json()


def _klartext(s):
    """HTML-Schnipsel aus den CMS-Feldern in reinen Text verwandeln."""
    return re.sub(r"\s+", " ", re.sub(r"<[^>]*>", " ", html.unescape(s or ""))).strip()


def _detail(sess, cid):
    """Titel, Datum und Dateien einer Sitzung über die CMS-Schnittstelle lesen.

    Die Detailseiten helfen hier nicht weiter: die Downloads werden erst im
    Browser nachgeladen, und der CDN liefert je nach URL eine abgespeckte
    Fassung ohne Dateiangaben. Die Schnittstelle /CMS/content ist dagegen
    einheitlich. Sie nennt im Feld data_widget_data unter dem Schlüssel
    "downloads" die contentids der angehängten Dateien; jede davon ist selbst
    ein Inhaltsobjekt mit Dateiname und Datei-ID."""
    j = _content(sess, cid)
    titel = _klartext(j.get("data_articleHeadline") or j.get("data_listlabel") or "") \
            or f"Sitzung {cid}"
    datum = (j.get("data_eventDate_start") or "").strip()   # DD.MM.YYYY

    datei_ids = []
    try:
        for section in json.loads(j.get("data_widget_data") or "[]"):
            for col in section.get("cols", []):
                for feld in col.get("fields", []):
                    for ct in (feld.get("data", {}) or {}).get("contenttypes", []):
                        if ct.get("key") == "downloads":
                            datei_ids += [str(x) for x in ct.get("contentids", [])]
    except Exception as e:
        print(f"     ! Widget-Daten von {cid} unlesbar: {e}")

    files, gesehen = [], set()
    for fid_content in dict.fromkeys(datei_ids):
        try:
            fj = _content(sess, fid_content)
        except Exception:
            continue
        meta = fj.get("data_filemeta") or ""
        if isinstance(meta, str):
            try:
                meta = json.loads(meta)
            except Exception:
                meta = {}
        name = (meta.get("originalname") or fj.get("data_file_name") or "").strip()
        fid = meta.get("fileid") or ""
        if not name or not fid or name.lower() in SEITEN_DATEIEN or fid in gesehen:
            continue
        gesehen.add(fid)
        files.append({"name": name, "href": f"{BASE}/CMS/get/file/{fid}"})
    return titel, datum, files


def discover_sessions():
    """Sitzungen inklusive Datei-Links einsammeln, ohne Browser.

    Frühere Fassungen lasen die Kacheln der Webseite aus. Die zeigt pro
    Jahres-Reiter aber nur 12 Einträge, wodurch die ersten Sitzungen eines
    Jahres systematisch fehlten. Der Listen-Endpoint liefert alle."""
    ids, _ = _session_ids()
    print(f"     {len(ids)} Sitzungskacheln über den Listen-Endpoint gefunden.", flush=True)

    # Index-Cache: schon bekannte Sitzungen werden nicht erneut abgefragt.
    # Das macht spätere Läufe schnell, weil nur neue Sitzungen dazukommen.
    cache = {}
    if INDEX_JSON.exists() and "--neu" not in sys.argv:
        try:
            cache = json.load(open(INDEX_JSON, encoding="utf-8"))
        except Exception:
            cache = {}
    offen = [c for c in ids if c not in cache]
    print(f"     {len(cache)} aus dem Index bekannt, {len(offen)} neu abzufragen.", flush=True)

    lokal = threading.local()

    def sitzung(cid):
        if not hasattr(lokal, "sess"):
            lokal.sess = requests.Session(); lokal.sess.headers.update(UA)
        titel, datum, files = _detail(lokal.sess, cid)
        return cid, {"titel": titel, "datum": datum, "files": files}

    if offen:
        fertig = 0
        with cf.ThreadPoolExecutor(max_workers=WORKERS) as pool:
            for fut in cf.as_completed([pool.submit(sitzung, c) for c in offen]):
                try:
                    cid, eintrag = fut.result()
                    cache[cid] = eintrag
                except Exception as e:
                    print(f"     ! Sitzung nicht ladbar: {e}", flush=True)
                fertig += 1
                if fertig % 25 == 0:
                    print(f"     ... {fertig}/{len(offen)} Sitzungen gelesen", flush=True)
        json.dump(cache, open(INDEX_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    sessions, ohne_dateien = [], []
    for cid in ids:
        e = cache.get(cid)
        if not e:
            continue
        datum = e.get("datum") or ""
        # Kein Jahresfilter an dieser Stelle: das CMS-Datum ist bei vielen
        # älteren Kacheln falsch gesetzt. Gefiltert wird später anhand des
        # Datums der Abstimmungsdatei.
        if not e["files"]:
            ohne_dateien.append(f"{e['titel']} ({datum or '?'})")
        sessions.append({"titel": e["titel"], "cid": cid, "datum": datum, "files": e["files"]})
    if ohne_dateien:
        print(f"     Hinweis: {len(ohne_dateien)} Sitzungen ohne hinterlegte Dateien, "
              f"z. B. {ohne_dateien[0]}", flush=True)
    return sessions


# ---------------------------------------------------------------------------
# SCHRITT 2: Relevante Excel-Dateien herunterladen
# ---------------------------------------------------------------------------
def _hole(url, ziel):
    """Datei laden, sofern noch nicht vorhanden. None bei defektem Link."""
    if ziel.exists():
        return ziel
    try:
        r = requests.get(url, timeout=90, headers=UA)
        r.raise_for_status()
    except Exception as e:
        print(f"     ! {ziel.name}: {e}")
        return None
    ziel.write_bytes(r.content)
    return ziel


def download_xlsx(sessions):
    """Lädt pro Sitzung die Abstimmungsergebnisse jeder Sitzungshälfte.

    Bevorzugt wird die xlsx. Ist deren Link defekt (kommt auf sh.ch vor) oder
    fehlt sie ganz, wird auf den gleichnamigen PDF-Report ausgewichen; der
    enthält dieselben Angaben und wird von pdf_report.py gelesen.
    Inhaltliche Duplikate (dieselbe Datei zweimal verlinkt) fallen über einen
    Hash des Dateiinhalts weg."""
    downloaded = []
    content_hashes = set()
    for s in sessions:
        # Kandidaten je Sitzungshälfte sammeln, Schlüssel = Dateiname ohne Endung
        haelften = {}
        for f in s["files"]:
            name = f["name"]
            nlow = name.lower()
            if not nlow.endswith((".xlsx", ".pdf")):
                continue
            if not any(m.search(name) for m in XLSX_MUSTER):
                continue
            if any(x in nlow for x in XLSX_AUSSCHLUSS):
                continue
            d = _datum_aus(name, s.get("titel", ""), s.get("datum", ""))
            if d and d[0] < AB_JAHR:
                continue
            stamm = re.sub(r"\.(xlsx|pdf)$", "", name, flags=re.I).strip().lower()
            # "Abst-2018-05.pdf" und "Abst_2018-05 (1).xlsx" gehören zusammen
            stamm = re.sub(r"\s*\(\d+\)$", "", stamm).replace("_", "-").strip()
            haelften.setdefault(stamm, {})[nlow[-4:]] = f

        for stamm, kand in sorted(haelften.items()):
            gewaehlt = None
            for typ, endung in (("xlsx", "xlsx"), ("pdf", ".pdf")):
                f = kand.get("xlsx") if typ == "xlsx" else kand.get(".pdf")
                if not f:
                    continue
                uid = hashlib.md5(f["href"].encode()).hexdigest()[:8]
                safe = re.sub(r"[^\w.-]", "_", f["name"])
                lokal = _hole(f["href"], OUT_DIR / f"{uid}_{safe}")
                if lokal:
                    gewaehlt = (typ, f["name"], lokal)
                    break
                print(f"     -> weiche für «{f['name']}» auf den PDF-Report aus")
            if not gewaehlt:
                continue
            typ, name, lokal = gewaehlt
            chash = hashlib.md5(lokal.read_bytes()).hexdigest()
            if chash in content_hashes:
                continue
            content_hashes.add(chash)
            downloaded.append({"titel": s["titel"], "cid": s.get("cid"),
                               "datum": s.get("datum", ""), "dateiname": name,
                               "pfad": str(lokal), "typ": typ})
    return downloaded


# ---------------------------------------------------------------------------
# SCHRITT 2b: Wortprotokolle (PDF) herunterladen
# ---------------------------------------------------------------------------
def download_protokolle(sessions, mit_pdf=False):
    """Lädt pro Sitzung die verlinkten Wortprotokoll-PDFs (Link-Text enthält
    'Protokoll'). Die Protokoll-Links tragen keinen sprechenden Dateinamen
    (get/file/<uuid>), darum wird nach cid und einem Kürzel der URL benannt.
    Inhaltlicher Dedup über einen Hash des Dateiinhalts.
    Rückgabe: dict cid -> Liste von {name, dateiname, url}."""
    by_cid = {}
    content_hashes = set()
    for s in sessions:
        cid = s.get("cid")
        for f in s["files"]:
            name = (f["name"] or "").strip()
            nlow = name.lower()
            ist_protokoll = (("protokoll" in nlow
                              or ("sitzung vom" in nlow and nlow.endswith(".pdf")))
                             and not any(x in nlow for x in PROT_EXCLUDE))
            if not ist_protokoll:
                continue
            uid = hashlib.md5(f["href"].encode()).hexdigest()[:8]
            local = PROT_DIR / f"protokoll_{cid}_{uid}.pdf"
            if not mit_pdf:
                # Standardfall: nur die Verweise festhalten. Die PDFs werden
                # vom Dashboard nicht gebraucht, und über alle Jahre wären es
                # mehrere hundert Megabyte.
                by_cid.setdefault(cid, []).append(
                    {"name": name, "dateiname": None, "url": f["href"]})
                continue
            if not local.exists():
                r = requests.get(f["href"], timeout=120, headers=UA)
                r.raise_for_status()
                if not r.content[:5].startswith(b"%PDF"):   # nur echte PDFs
                    continue
                local.write_bytes(r.content)
            chash = hashlib.md5(local.read_bytes()).hexdigest()
            if chash in content_hashes:
                continue
            content_hashes.add(chash)
            by_cid.setdefault(cid, []).append(
                {"name": name, "dateiname": local.name, "url": f["href"]})
    return by_cid


# ---------------------------------------------------------------------------
# SCHRITT 3: Eine Excel-Datei in strukturiertes JSON parsen
# ---------------------------------------------------------------------------
def parse_xlsx(path, sitzung_label, cid=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Spaltenlage aus der Kopfzeile bestimmen. Ältere Dateien führen vorne
    # zusätzlich "S/N Keypad" und "Nr.", die Namensspalten sitzen dann zwei
    # Spalten weiter rechts. Fest verdrahtete Spaltennummern lieferten dort
    # Keypad-IDs statt Namen.
    # Die Kopfzeile steht meist auf Zeile 1, in einzelnen Dateien eine Zeile
    # tiefer. Darum die ersten Zeilen nach "Nachnamen" absuchen.
    kopf_zeile = 1
    for r in range(1, 6):
        if any(isinstance(ws.cell(row=r, column=c).value, str)
               and str(ws.cell(row=r, column=c).value).strip().startswith("Nachname")
               for c in range(1, ws.max_column + 1)):
            kopf_zeile = r
            break

    spalte, abst_spalten = {}, []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=kopf_zeile, column=c).value
        if not isinstance(v, str):
            continue
        v = v.strip()
        if v.startswith("Abst."):
            abst_spalten.append(c)
        elif v.startswith("Nachname"):
            spalte["nachname"] = c
        elif v.startswith("Vorname"):
            spalte["vorname"] = c
        elif v.startswith("Fraktion"):
            spalte["fraktion"] = c
        elif v.startswith("Partei"):
            spalte["partei"] = c
    n_votes = len(abst_spalten)
    c_nach = spalte.get("nachname", 1)
    # Nur vorhandene Spalten verwenden. Einzelne ältere Dateien führen keine
    # Parteispalte; ein Rateversuch landete dort auf der ersten Stimmspalte und
    # trug "Ja"/"Nein" als Partei ein.
    c_vor = spalte.get("vorname")
    c_frak = spalte.get("fraktion")
    c_par = spalte.get("partei")

    # Summenzeilen unter der Mitgliederliste
    SUMMEN = {"ja", "nein", "enth", "enthaltung", "total", "v/a/n",
              "vakanz, abwesenheit, nicht-teilnahme"}

    members = []
    leer = 0
    for r in range(kopf_zeile + 1, ws.max_row + 1):
        nach = ws.cell(row=r, column=c_nach).value
        if nach is None or not str(nach).strip():
            leer += 1
            if leer >= 2:          # zwei leere Zeilen: Mitgliederliste zu Ende
                break
            continue
        leer = 0
        if str(nach).strip().lower() in SUMMEN:
            break
        frak = ws.cell(row=r, column=c_frak).value if c_frak else ""
        if frak in (None, "Fraktionen"):
            continue
        votes = [ws.cell(row=r, column=c).value for c in abst_spalten]
        fraktion = str(frak or "").strip()
        partei = str(ws.cell(row=r, column=c_par).value or "").strip() if c_par else ""
        members.append({
            "nachname": str(nach).strip(),
            "vorname": str(ws.cell(row=r, column=c_vor).value or "").strip() if c_vor else "",
            "fraktion": fraktion,
            "partei": partei or fraktion,       # ohne Parteispalte gilt die Fraktion
            "votes": votes,
        })

    def txt(r, c):
        return ws.cell(row=r, column=c).value

    # Metablock: Kopfzeile über "Traktandum" finden. Auch dieser Block ist in
    # älteren Dateien nach rechts verschoben, darum werden alle Spalten aus der
    # gefundenen Kopfzeile abgeleitet statt fest verdrahtet.
    # Erkannt wird die Kopfzeile an "Betreff" und "Abstimmung": die Spalte mit
    # dem Traktandumstext heisst je nach Datei "Traktandum", "Traktanden" oder
    # trägt gleich einen Hinweistext statt eines Titels.
    # Kopfzeile des Metablocks suchen. Sie steht nach der Mitgliedertabelle und
    # ist in manchen Dateien über mehrere Zeilen verteilt; die Spalte mit dem
    # Traktandumstext heisst mal "Traktandum", mal "Traktanden" und trägt
    # gelegentlich statt eines Spaltennamens gleich einen Hinweistext.
    meta_start = None
    for r in range(2, ws.max_row + 1):
        zeile = [v.strip() for v in (txt(r, c) for c in range(1, ws.max_column + 1))
                 if isinstance(v, str)]
        if "Betreff" in zeile or "Abstimmung" in zeile:
            meta_start = r
            break

    col_nr, col_trakt, col_typ, col_inv = None, None, 10, 12      # Rückfallwerte
    col_typ_gefunden = col_inv_gefunden = False
    if meta_start:
        # Fenster auch zwei Zeilen nach oben: "Nr." steht mitunter eine Zeile
        # über "Betreff" und "Abstimmung".
        for r in range(max(2, meta_start - 2), min(meta_start + 6, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                h = txt(r, c)
                if not isinstance(h, str):
                    continue
                h = h.strip()
                # Nur den ersten Treffer nehmen: weiter unten stehen Inhalte
                # wie "Traktandenliste", die sonst die Spalte verschieben.
                if h.rstrip(".") == "Nr" and col_nr is None:
                    col_nr = c
                elif h in ("Traktandum", "Traktanden") and col_trakt is None:
                    col_trakt = c
                elif h == "Betreff" and col_typ_gefunden is False:
                    col_typ, col_typ_gefunden = c, True
                elif h == "Abstimmung" and col_inv_gefunden is False:
                    col_inv, col_inv_gefunden = c, True
    # Fehlt eine der beiden Spalten in der Kopfzeile, aus der anderen ableiten:
    # die Nummernspalte steht stets unmittelbar links vom Traktandumstext.
    if col_nr is None and col_trakt is not None:
        col_nr = col_trakt - 1
    if col_nr is None:
        col_nr = 1
    if col_trakt is None:
        col_trakt = col_nr + 1

    # Vorspann eines Geschäfts, z. B. "Die Abstimmungen 1 bis 5 beziehen sich
    # auf folgendes Geschäft:". Hinweise wie "Die Abstimmung erfolgt mittels
    # Namensaufruf" dürfen nicht mitgehen.
    VORSPANN = re.compile(r"^Die Abstimmung(?:en)?\b.{0,40}?bezieh(?:t|en)\s+sich\s+auf", re.S)

    geschaeft_map = {}
    if meta_start:
        for r in range(meta_start, ws.max_row + 1):
            c2 = txt(r, col_trakt)
            if isinstance(c2, str) and VORSPANN.match(c2):
                buf, rr = [c2], r + 1
                while rr <= ws.max_row:
                    n1, n2 = txt(rr, col_nr), txt(rr, col_trakt)
                    if isinstance(n1, str) and re.match(r"Abstimmung\s+\d+", n1):
                        break
                    if isinstance(n2, str) and VORSPANN.match(n2):
                        break
                    if isinstance(n2, str) and n2.strip():
                        buf.append(n2)
                    rr += 1
                g = " ".join(buf)
                # technischen Vorspann entfernen: "Die Abstimmungen Nr. X-Y
                # beziehen sich auf folgendes Geschäft:" -> nur das Geschäft behalten
                g = re.sub(r"^Die Abstimmung(?:en)?\b.*?bezieh(?:t|en)\s+sich\s+auf\s+"
                           r"(?:folgendes\s+Gesch[äa]ft:?\s*|(?:den|das|die)\s+)?",
                           "", g, flags=re.S).strip()
                m = re.search(r"(?:Nr\.\s*)?(\d+)\s*(?:(?:-|bis)\s*(\d+))?", c2)
                if m:
                    a = int(m.group(1))
                    b = int(m.group(2)) if m.group(2) else a
                    for n in range(a, b + 1):
                        geschaeft_map[n] = g

    votes_meta = []
    if meta_start:
        for r in range(meta_start, ws.max_row + 1):
            c1 = txt(r, col_nr)
            if isinstance(c1, str) and re.match(r"Abstimmung\s+\d+", c1):
                nr = int(re.search(r"\d+", c1).group())
                titel = txt(r, col_trakt) or ""
                # Der Betreff steht mal auf der Zeile der Abstimmung, mal auf
                # einer der Folgezeilen -> die ersten Zeilen des Blocks absuchen.
                typ = ""
                for rr in range(r, r + 5):
                    cand = txt(rr, col_typ)
                    if isinstance(cand, str) and cand.strip():
                        typ = cand
                        break
                details, inverted = [], None
                for rr in range(r + 1, r + 14):
                    t2, t12, c1n = txt(rr, col_trakt), txt(rr, col_inv), txt(rr, col_nr)
                    if isinstance(c1n, str) and re.match(r"Abstimmung\s+\d+", c1n):
                        break
                    if isinstance(t2, str) and t2.startswith("Die Abstimmung"):
                        break
                    if isinstance(t2, str) and t2.strip() and not t2.startswith(
                        ("Ja bedeutet", "Nein bedeutet", "fakultatives")):
                        details.append(t2.strip())
                    if isinstance(t12, str) and t12.startswith("Ja bedeutet"):
                        # Der erklärende Text kann über mehrere Spalten verteilt
                        # sein (z.B. "Ja bedeutet" | "Zustimmung Kommission").
                        # Darum die Zellen der Zeile zusammenziehen.
                        extra = [str(txt(rr, cc)).strip()
                                 for cc in range(col_inv, col_inv + 6)
                                 if isinstance(txt(rr, cc), str) and txt(rr, cc).strip()]
                        inverted = " ".join(extra).strip()
                votes_meta.append({
                    "nr": nr, "titel": str(titel).strip(), "typ": str(typ).strip(),
                    "details": " ".join(details), "inverted_note": inverted,
                    "geschaeft": geschaeft_map.get(nr, ""),
                })
    votes_meta.sort(key=lambda x: x["nr"])
    # Die Metadaten-Liste kann mehr Abstimmungen beschreiben als es Stimmspalten
    # gibt (z.B. wenn die Traktanden-Nummerierung über Vormittag/Nachmittag/Abend
    # durchläuft, die Datei aber nur eine Tageshälfte enthält). Die tatsächlichen
    # Stimmspalten (n_votes) sind massgeblich: auf die ersten n_votes begrenzen.
    votes_meta = votes_meta[:n_votes]
    url = DETAIL_URL_TMPL.format(cid=cid) if cid else None
    return {"sitzung": sitzung_label, "n_votes": n_votes, "cid": cid, "url": url,
            "members": members, "votes": votes_meta, "quelle": Path(path).name}


# ---------------------------------------------------------------------------
# SCHRITT 4: Sitzungslabel aus Dateiname ableiten (Datum) + zusammenführen
# ---------------------------------------------------------------------------
def _datum_tuple(s):
    """(Jahr, Monat, Tag) aus dem Sitzungslabel."""
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s["sitzung"])
    return (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else (0, 0, 0)


def _datum_aus(dateiname, titel="", cms_datum=""):
    """Sitzungsdatum als (Jahr, Monat, Tag) bestimmen.

    Reihenfolge: Datum im Dateinamen, sonst Datum im Sitzungstitel
    ("7. Kantonsratssitzung vom 14.05.2018"), sonst das Datum aus dem CMS."""
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", dateiname or "")
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return (y, mo, d)
        if 1 <= d <= 12 and 1 <= mo <= 31:      # vertauscht, z.B. "20213008"
            return (y, d, mo)
    for quelle in (titel, cms_datum):
        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(20\d{2})", quelle or "")
        if m:
            return (int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def label_from(titel, dateiname, sitzungsdatum=""):
    """Kombiniert Sitzungstitel mit Datum und Tageszeit, damit Vormittag,
    Nachmittag, Abend und Doppelsitzungen unterscheidbar bleiben."""
    # ältere Titel tragen das Datum schon im Text ("7. Kantonsratssitzung vom
    # 14.05.2018"), das käme sonst doppelt vor
    titel = re.sub(r"\s+vom\s+\d{1,2}\.\d{1,2}\.20\d{2}\.?\s*$", "", titel).strip()
    dl = dateiname.lower()
    half = ""
    if "vormittag" in dl:
        half = " (Vormittag)"
    elif "nachmittag" in dl:
        half = " (Nachmittag)"
    elif "abend" in dl:
        half = " (Abend)"
    # ältere Namen ohne Tageszeit: "…Teil 1…" bzw. "Abst_2018-07" (Sitzungsnummer)
    mt = re.search(r"teil\s*(\d)", dl)
    mn = re.search(r"abst[_\s-]?20\d{2}[-_](\d{1,2})", dl)
    if mt:
        half += f" (Teil {mt.group(1)})" if half else f" (Teil {mt.group(1)})"
    elif not half and mn:
        half = f" (Sitzung {int(mn.group(1))})"
    # mehrteilige Nachmittage: "…Nachmittag1.xlsx" / "…Nachmittag 2. Teil…"
    mz = re.search(r"(?:vormittag|nachmittag|abend)\s*(\d)\b", dl)
    if mz and "Teil" not in half:
        half = half[:-1] + f" {mz.group(1)})"

    d = _datum_aus(dateiname, titel, sitzungsdatum)
    datum = f" · {d[2]:02d}.{d[1]:02d}.{d[0]}" if d else ""
    return f"{titel}{datum}{half}"


def main():
    print("1/4  Sitzungen suchen ...")
    sessions = discover_sessions()
    print(f"     {len(sessions)} Sitzungskacheln gefunden.")

    print("2/4  Excel-Dateien herunterladen ...")
    files = download_xlsx(sessions)
    print(f"     {len(files)} Abstimmungs-Excel geladen.")
    if not files:
        print("     KEINE passenden Dateien. Marker prüfen:", XLSX_MARKER)
        sys.exit(1)

    print("3/4  Parsen ...")
    parsed = []
    # Erst alle xlsx, danach die PDF-Ausweichfälle: dann steht für das PDF die
    # Mitgliederliste der anderen Sitzungshälfte als Namensraster bereit.
    for durchgang in ("xlsx", "pdf"):
        for f in [x for x in files if x.get("typ", "xlsx") == durchgang]:
            label = label_from(f["titel"], f["dateiname"], f.get("datum", ""))
            try:
                if durchgang == "xlsx":
                    s = parse_xlsx(f["pfad"], label, f.get("cid"))
                else:
                    from pdf_report import parse_pdf
                    roster = next((p["members"] for p in parsed
                                   if p.get("cid") == f.get("cid")), None)
                    s = parse_pdf(f["pfad"], label, f.get("cid"), roster)
                if not s["members"] or not s["n_votes"]:
                    print(f"     ✗ {label}: keine Stimmen erkannt, übersprungen")
                    continue
                if any(p["sitzung"] == s["sitzung"] for p in parsed):
                    print(f"     · {label}: bereits erfasst, übersprungen")
                    continue
                parsed.append(s)
                quelle = "" if durchgang == "xlsx" else "  [aus PDF-Report]"
                print(f"     ✓ {label}  ({s['n_votes']} Abst., "
                      f"{len(s['members'])} Mitgl.){quelle}")
            except Exception as e:
                print(f"     ✗ Fehler bei {f['dateiname']}: {e}")

    mit_pdf = "--protokolle" in sys.argv        # PDFs wirklich herunterladen
    print(f"     Wortprotokolle {'herunterladen' if mit_pdf else 'verlinken'} ...")
    prot_by_cid = download_protokolle(sessions, mit_pdf=mit_pdf)
    for s in parsed:
        s["protokolle"] = prot_by_cid.get(s.get("cid"), [])
    n_prot = sum(len(v) for v in prot_by_cid.values())
    print(f"     {n_prot} Wortprotokolle für {len(prot_by_cid)} Sitzungstage.")

    # Neueste zuerst: echtes Datum aus dem Label ziehen (DD.MM.YYYY),
    # Vormittag vor Nachmittag. Fällt auf 0 zurück, wenn kein Datum gefunden.
    def sort_key(s):
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s["sitzung"])
        datum = (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else (0, 0, 0)
        halb = 0 if "Vormittag" in s["sitzung"] else 1
        return (datum, halb)
    parsed.sort(key=sort_key, reverse=True)

    # --- Legislatur-Zuordnung ---
    # Die Amtsdauer des Kantonsrats beträgt vier Jahre und beginnt jeweils am
    # 1. Januar nach den Gesamterneuerungswahlen im September davor. Die Grenzen
    # werden darum fest gesetzt statt aus dem Mitgliederwechsel geschätzt: die
    # Schätzung verschob den Beginn auf die erste Sitzung, für die Daten
    # vorliegen, und übersah Wechsel mit wenig Fluktuation.
    chrono = list(reversed(parsed))  # älteste zuerst

    def leg_von_datum(y, m, d):
        for nr, (start, ende) in LEGISLATUREN.items():
            if start <= (y, m, d) <= ende:
                return nr
        return None

    unbekannt = []
    prev_names, leg_fallback = None, max(LEGISLATUREN) + 1
    for s in chrono:
        y, m, d = _datum_tuple(s)
        nr = leg_von_datum(y, m, d)
        if nr is None:
            # ausserhalb der bekannten Amtsdauern: Wechsel wieder schätzen
            names = {(mm["nachname"], mm["vorname"]) for mm in s["members"]}
            if prev_names is not None and len(names - prev_names) > 15:
                leg_fallback += 1
            nr = leg_fallback
            unbekannt.append(s["sitzung"])
            prev_names = names
        s["legislatur"] = nr
    if unbekannt:
        print(f"     Hinweis: {len(unbekannt)} Sitzungen ausserhalb der "
              f"hinterlegten Amtsdauern, z. B. {unbekannt[0]}")

    # Zeitraum-Label + aktuelle Mitglieder je Legislatur
    from collections import defaultdict
    by_leg = defaultdict(list)
    for s in chrono:
        by_leg[s["legislatur"]].append(s)

    def datum_tuple(s):
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s["sitzung"])
        return (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else (0, 0, 0)

    leg_meta = {}
    for leg, ss in by_leg.items():
        ss_sorted = sorted(ss, key=datum_tuple)
        d0, d1 = datum_tuple(ss_sorted[0]), datum_tuple(ss_sorted[-1])
        # aktive Mitglieder dieser Legislatur = Mitglieder der neuesten Sitzung
        newest = ss_sorted[-1]
        aktiv = [f"{m['nachname']}|{m['vorname']}" for m in newest["members"]]
        amt = LEGISLATUREN.get(leg)
        # Amtsdauer als Label, nicht die erste Sitzung mit Daten: sonst stünde
        # dort ein zufälliges Datum aus der Mitte der Amtszeit.
        if amt:
            label = f"Legislatur {amt[0][0]}–{amt[1][0]}"
            von = f"{amt[0][2]:02d}-{amt[0][1]:02d}-{amt[0][0]}"
            bis = f"{amt[1][2]:02d}-{amt[1][1]:02d}-{amt[1][0]}"
        else:
            label = f"Legislatur ab {d0[2]:02d}.{d0[1]:02d}.{d0[0]}"
            von = f"{d0[2]:02d}-{d0[1]:02d}-{d0[0]}"
            bis = f"{d1[2]:02d}-{d1[1]:02d}-{d1[0]}"
        leg_meta[leg] = {
            "nummer": leg,
            "von": von,
            "bis": bis,
            "label": label,
            "erste_sitzung": f"{d0[2]:02d}.{d0[1]:02d}.{d0[0]}",
            "letzte_sitzung": f"{d1[2]:02d}.{d1[1]:02d}.{d1[0]}",
            "aktive_mitglieder": aktiv,
            "n_sitzungen": len(ss),
        }

    # höchste Legislaturnummer = aktuelle
    aktuelle_leg = max(leg_meta) if leg_meta else 1

    print("4/4  Schreiben ...")
    json.dump({"sessions": parsed, "n_sessions": len(parsed),
               "legislaturen": leg_meta, "aktuelle_legislatur": aktuelle_leg},
              open(JSON_OUT, "w"), ensure_ascii=False, indent=1)
    total_votes = sum(s["n_votes"] for s in parsed)
    print(f"     {JSON_OUT}: {len(parsed)} Sitzungen, {total_votes} Abstimmungen.")


if __name__ == "__main__":
    main()
